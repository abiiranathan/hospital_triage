from tkinter import *
from tkinter import ttk
from collections import OrderedDict
from mysql.connector import connect
import openpyxl
from cashtree import Treeview
import Pmw
from widgets import SearchEngine, Treeview
from tkinter.simpledialog import askstring
import os


class Billing:
    def __init__(self,parent, num_rows=4, num_cols=4):
        self.parent=parent
        self.num_rows= num_rows
        self.num_cols= num_cols
        self.row_in_focus=1
        self.column_in_focus=0
        self.current_row = 0

        self.create_table(num_rows, num_cols)
        self.parent.bind_all("<Up>", self.up)
        self.parent.bind_all("<Down>", self.down)
        self.parent.bind_all("<Right>", self.right)
        self.parent.bind_all("<Left>", self.left)

    def create_table(self, num_rows, num_cols):
        for i in range(num_rows):
            for j in range(num_cols):
                ent = ttk.Entry(self.parent)
                ent.configure(font="Consolas 12")
                ent.grid(row=i, column=j, sticky='nsew')
                self.parent.grid_columnconfigure(i, weight=1)
                self.parent.grid_rowconfigure(j, weight=1)

        for index, entry in enumerate(self.parent.children.values()):
            self.header_rows=OrderedDict()

            if isinstance(entry, Entry):
                info= entry.grid_info()

                istcell=self.find_entry(self.row_in_focus, self.column_in_focus)
                istcell.focus_set()

                if info['row']==0:
                    entry.configure(font="Arial 12 bold")
                    self.header_rows[info['column']]=entry

                if info['column']==0:
                    entry.configure(width=10)

                if info['column']==1:
                    entry.configure(width=60)

                if info['column']==2:
                    entry.configure(width=20)

                if info['column']==3:
                    entry.configure(width=20)

                if info['column']==4:
                    entry.configure(width=25)

            # Insert header items
            self.header_labels=["Item ID","Item", "Qty", "Unit Cost", "SubTotal"]
            for key, value in self.header_rows.items():
                value.insert(0, self.header_labels[key])
                value.configure(state=DISABLED, foreground='blue')

    def find_entry(self, row, column,find_header=False):
        for child in self.parent.children.values():
            info = child.grid_info()
            if int(info['row'])==int(row) and int(info['column'])==int(column):
                if row == 0:
                    if not find_header:
                        return None
                    else:
                        return child
                else:
                    return child

    def find_in_grid(self, row, column, get_header=False):
        for child in self.parent.children.values():
            info = child.grid_info()

            if isinstance(child, Entry):
                if int(info['row'])==int(row) and int(info['column'])==int(column):
                    if row == 0:
                        if get_header:
                            return child.get()
                        return None
                    else:
                        try:
                            return float(child.get())
                        except ValueError:
                            return None

    def find_cell(self, row, column, get_header=False):
        for child in self.parent.children.values():
            info = child.grid_info()

            if isinstance(child, Entry):
                if int(info['row'])==int(row) and int(info['column'])==int(column):
                    if row == 0:
                        if get_header:
                            return child.get()
                        return None

                    else:
                        try:
                            return float(child.get().replace(", ", "")) # Convert strings to Numbers
                        except:
                            return child.get()

    def get_row_data(self, row_num):
        data=[]
        column=0
        while column<self.num_cols:
            value = self.find_cell(row_num, column,get_header=True)
            data.append(value)
            column +=1
        return data

    def submit_payment(self):
        data= self.findall()

    def findall(self):
        data_lists=[]
        for i in range(0,self.num_rows):
            data=self.get_row_data(i)
            data_lists.append(tuple(data))

        self.export_to_excel(data_lists)
        return data_lists

    def export_to_excel(self,output_data):
        output_workbook =openpyxl.Workbook()
        output_sheet = output_workbook.create_sheet("Billing Sheet", 0)

        for row in output_data:
            rowIn = output_data.index(row)
            for col in range(len(output_data[0])):
                output_sheet.cell(row=rowIn+1, column=col+1).value=output_data[rowIn][col]

        try:
            output_workbook.save('billing.xlsx')
            os.startfile("billing.xlsx")
        except IOError:
            pass

    def clear(self):
        for child in self.parent.children.values():
            if isinstance(child, Entry):
                child.delete(0, END)

        self.current_row=0

    def compute(self, grandtotal_entry):
        subtotals=[]

        for index, child in enumerate(self.parent.children.values()):
            qty= self.find_in_grid(row=index, column=2)
            unit=self.find_in_grid(row=index, column=3)
            subtotal = self.find_entry(row=index, column=4)
            grand_total_ent= self.find_entry(row=self.num_rows, column=self.num_cols-1)

            if qty and unit and subtotal:
                subt = qty*unit
                subtotal.delete(0, END)
                subtotals.append(subt)
                subtotal.insert(0, self.sep(subt))

            if (not qty or not unit) and subtotal:
                if subtotal.get() !="":
                    subtotal.delete(0,END)

        # Cumpute Grand total
        grandtotal= self.sep(str(sum(subtotals)))
        self.grandtotal_entry = grandtotal_entry
        self.grandtotal_entry.delete(0, END)
        self.grandtotal_entry.insert(0, grandtotal +" /=")


    def insert_item(self, item_no, item, cost):
        self.current_row  +=1

        row = self.current_row
        idvalue = self.find_cell(row=row, column=0)
        itemvalue = self.find_cell(row=row, column=1)
        pricevalue =self.find_cell(row=row, column=3)

        if idvalue and itemvalue and pricevalue:
            self.current_row +=1
            row +=1

        idcell = self.find_entry(row=row, column=0)
        itemcell = self.find_entry(row=row, column=1)
        pricecell =self.find_entry(row=row, column=3)

        idcell.delete(0,END)
        idcell.insert(0, item_no)

        itemcell.delete(0,END)
        itemcell.insert(0, item)

        pricecell.delete(0,END)
        pricecell.insert(0, int(str(cost).replace(",","")))

    def sep(self, s, thou=", ", dec="."):
        try:
            '''Generates thousand seperators'''
            if str(s).isdigit():
                s = str(s) + '.0'
            integer, decimal = str(s).split(".")
            integer = re.sub(r"\B(?=(?:\d{3})+$)", thou, integer)
            return integer + dec + decimal
        except:
            return ""

    def up(self, event):
        info=event.widget.grid_info()
        current_row, current_col= info['row'], info['column']
        upper_row, upper_col=current_row-1, current_col
        cell = self.find_entry(upper_row, upper_col)
        if cell:
            cell.focus_set()

    def down(self, event):
        info=event.widget.grid_info()
        current_row, current_col= info['row'], info['column']
        lower_row, lower_col=current_row+1, current_col
        cell = self.find_entry(lower_row, lower_col)
        if cell:
            cell.focus_set()

    def right(self, event):
        info=event.widget.grid_info()
        current_row, current_col= info['row'], info['column']
        right_row, right_col=current_row, current_col +1
        cell = self.find_entry(right_row, right_col)
        if cell:
            cell.focus_set()

    def left(self, event):
        info=event.widget.grid_info()
        current_row, current_col= info['row'], info['column']
        left_row, left_col=current_row, current_col -1
        cell = self.find_entry(left_row, left_col)

        if cell:
            cell.focus_set()

    def insert_row(self):
        for index in range(self.num_cols):
            ent = ttk.Entry(self.parent, width=10)
            ent.grid_columnconfigure(index, weight=1)
            ent.configure(font="Consolas 12")
            ent.grid(row= self.num_rows+1, column=index, sticky='ew')

        self.num_rows +=1


    def insert_column(self):
        columnname= askstring("Column?", "What's the column name?")
        if columnname:
            for index in range(self.num_rows):
                ent = ttk.Entry(self.parent, width=20)
                ent.grid_columnconfigure(index, weight=1)
                ent.configure(font="Consolas 12")
                ent.grid(column= self.num_cols+1, row=index, sticky='ew')

            self.num_cols +=1

            top = self.find_entry(row=0, column=self.num_cols, find_header=1)
            top.insert(0, columnname)
            top.configure(state=DISABLED, foreground='blue', font="Arial 12 bold")

class Spreadshets(Pmw.ScrolledFrame):
    '''Spread sheets class implemets a Billing class
    The billing class is an excel like implementation
    self.sheet = Billing(parent, rows, columns)
    '''
    def __init__(self,parent, rows, columns):
        Pmw.ScrolledFrame.__init__(self,
            usehullsize=1,
            hull_width=1300,
            hull_height=600)

        self.parent=parent
        style = ttk.Style()
        # style.theme_use("clam")
        style.configure("TButton",borderwidth=8, background='powderblue',
            font="Arial 14 bold", foreground='green')
        self.parent.geometry("{}x{}+0+0".format(
            self.parent.winfo_screenwidth(),
            self.parent.winfo_screenheight()))

        sheetframe= Frame(self.interior())
        self.sheet = Billing(sheetframe, rows, columns)

        btnfrm= Frame(self.interior())
        self.top = Frame(self.interior())

        search_engine= SearchEngine(self.top, command_to_run=self.insert_item)
        search_engine.connect_to(user="root", database="patients", password="cmaster2018")
        search_engine.use(table="hospital_items")
        search_engine.set_width(100)
        search_engine.set_padding()

        ttk.Button(btnfrm, text="Submit",
            command=self.sheet.submit_payment ).grid(row=0, column=0,padx=2)

        ttk.Button(btnfrm, text="Export to Excel",
            command=self.sheet.findall ).grid(row=0, column=1,padx=2)
        ttk.Button(btnfrm, text="CLEAR",
            command=self.sheet.clear).grid(row=0,column=2,padx=2)
        ttk.Button(btnfrm, text="COMPUTE TOTAL",
            command=lambda: self.sheet.compute(self.grandTotal)).grid(row=0,column=3,padx=2)
        ttk.Button(btnfrm, text="INSERT ROW",
            command=self.sheet.insert_row).grid(row=0,column=4,padx=2)

        tot_frm= Frame(self.interior())

        self.grandTotal=ttk.Entry(tot_frm, width=40)
        self.grandTotal.pack(side=RIGHT,anchor='e')
        self.grandTotal.configure(font="Arial 14 bold")

        Label(tot_frm, text="GRAND TOTAL: ",
            font="Calibri 14 bold").pack(side=RIGHT, anchor='w')

        tot_frm.pack(expand=1, fill=X, anchor='ne',pady=4)
        btnfrm.pack(pady=5)
        self.top.pack(expand=1, fill=BOTH)
        sheetframe.pack(expand=1, fill=BOTH, padx=10)
        self.pack(expand=1, fill=BOTH,padx=10,pady=10)


    def insert_item(self, selection):
        ''' self.insert_item(selection)
        selection shouls be an iterable, list or tuple
        '''
        item_no, item, price = selection[0],selection[1],selection[2]
        self.sheet.insert_item(item_no, item, price)

if __name__ == '__main__':
    root=Tk()
    app= Spreadshets(root, 20,5)
    root.mainloop()
